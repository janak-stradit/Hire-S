"""Audit all HireX-created workbooks for structural and data-quality anomalies."""

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")


def workbook_paths() -> list[Path]:
    return sorted(
        path for path in ROOT.rglob("*.xlsx")
        if "venv" not in path.parts and not path.name.startswith("~$")
    )


def audit_workbook(path: Path) -> dict[str, Any]:
    relative = path.relative_to(ROOT)
    workbook = load_workbook(path, read_only=False, data_only=False)
    issues: list[str] = []
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if any(value is None or not str(value).strip() for value in headers):
            issues.append(f"{sheet.title}: blank header")
        if len(headers) != len(set(headers)):
            issues.append(f"{sheet.title}: duplicate header")
        formula_errors = []
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and any(error in cell.value for error in FORMULA_ERRORS):
                    formula_errors.append(cell.coordinate)
        if formula_errors:
            issues.append(f"{sheet.title}: formula errors at {formula_errors[:10]}")
        sheets.append({
            "name": sheet.title,
            "data_rows": max(sheet.max_row - 1, 0),
            "columns": sheet.max_column,
            "freeze_panes": str(sheet.freeze_panes or ""),
            "tables": {name: sheet.tables[name].ref for name in sheet.tables},
        })
        _audit_known_schema(relative, sheet, headers, issues)
    workbook.close()
    return {"path": str(relative), "status": "PASS" if not issues else "FAIL", "issues": issues, "sheets": sheets}


def _audit_known_schema(path: Path, sheet, headers: list[Any], issues: list[str]) -> None:
    name = path.name
    if name == "hirex_candidate_master_with_extracted_resumes.xlsx":
        _require_schema(sheet, 43, issues)
        _duplicates(sheet, headers, ["candidate_id", "resume_id", "parsed_resume_id", "email", "phone"], issues)
        _resume_integrity(sheet, headers, issues)
    elif name == "hirex_candidates.xlsx":
        _require_schema(sheet, 47, issues)
        _duplicates(sheet, headers, ["candidate_id", "resume_id", "parsed_resume_id", "email", "phone"], issues)
        _resume_integrity(sheet, headers, issues)
        expected = {
            "source_type": "resume_extracted",
            "agent_processing_allowed": True,
        }
        for header, value in expected.items():
            column = headers.index(header) + 1
            invalid = [row for row in range(2, sheet.max_row + 1) if sheet.cell(row, column).value != value]
            if invalid:
                issues.append(f"{header}: unexpected values in rows {invalid[:10]}")
        path_column = headers.index("storage_path") + 1
        missing_files = [
            row for row in range(2, sheet.max_row + 1)
            if not Path(str(sheet.cell(row, path_column).value)).exists()
        ]
        if missing_files:
            issues.append(f"storage_path: missing source files in rows {missing_files[:10]}")
    elif name == "test_synthetic_candidates.xlsx":
        _require_schema(sheet, 47, issues)
        _duplicates(sheet, headers, ["candidate_id", "resume_id", "parsed_resume_id", "email", "phone"], issues)
        _resume_integrity(sheet, headers, issues)
        expected = {"source_type": "synthetic", "agent_processing_allowed": False}
        for header, value in expected.items():
            column = headers.index(header) + 1
            invalid = [row for row in range(2, sheet.max_row + 1) if sheet.cell(row, column).value != value]
            if invalid:
                issues.append(f"{header}: unexpected values in rows {invalid[:10]}")
        for header in ("skills", "education", "certifications", "projects", "summary", "raw_text", "resume"):
            column = headers.index(header) + 1
            blanks = [row for row in range(2, sheet.max_row + 1) if not sheet.cell(row, column).value]
            if blanks:
                issues.append(f"{header}: blank values in rows {blanks[:10]}")
    elif name == "jd_input.xlsx":
        _require_schema(sheet, 19, issues)
        _duplicates(sheet, headers, ["requirement_id"], issues)
        status_column = headers.index("status") + 1
        statuses = [str(sheet.cell(row, status_column).value or "").lower() for row in range(2, sheet.max_row + 1)]
        if statuses.count("active") != 1:
            issues.append(f"status: expected one active requirement, found {statuses.count('active')}")
        for row in range(2, sheet.max_row + 1):
            minimum = sheet.cell(row, headers.index("experience_min") + 1).value
            maximum = sheet.cell(row, headers.index("experience_max") + 1).value
            salary_min = sheet.cell(row, headers.index("salary_min") + 1).value
            salary_max = sheet.cell(row, headers.index("salary_max") + 1).value
            pass_score = sheet.cell(row, headers.index("screening_pass_score") + 1).value
            review_score = sheet.cell(row, headers.index("screening_review_score") + 1).value
            created = sheet.cell(row, headers.index("created_at") + 1)
            if maximum < minimum:
                issues.append(f"row {row}: invalid experience range")
            if salary_max < salary_min:
                issues.append(f"row {row}: invalid salary range")
            if not 0 <= review_score < pass_score <= 100:
                issues.append(f"row {row}: invalid screening thresholds")
            if not isinstance(created.value, str) or not created.value.endswith("Z") or created.number_format != "@":
                issues.append(f"row {row}: created_at must be ISO-8601 text")


def _require_schema(sheet, columns: int, issues: list[str]) -> None:
    if sheet.max_row < 2:
        issues.append("shape: workbook contains no data rows")
    if sheet.max_column != columns:
        issues.append(f"shape: expected {columns} columns, found {sheet.max_column}")


def _duplicates(sheet, headers: list[Any], unique_columns: list[str], issues: list[str]) -> None:
    for header in unique_columns:
        column = headers.index(header) + 1
        values = [sheet.cell(row, column).value for row in range(2, sheet.max_row + 1)]
        nonblank = [str(value).strip().casefold() for value in values if value not in (None, "")]
        duplicates = len(nonblank) - len(set(nonblank))
        if duplicates:
            issues.append(f"{header}: {duplicates} duplicate values")


def _resume_integrity(sheet, headers: list[Any], issues: list[str]) -> None:
    raw_column = headers.index("raw_text") + 1
    resume_column = headers.index("resume") + 1
    mismatches = []
    blanks = []
    for row in range(2, sheet.max_row + 1):
        raw = sheet.cell(row, raw_column).value
        resume = sheet.cell(row, resume_column).value
        if not raw or not resume:
            blanks.append(row)
        elif raw != resume:
            mismatches.append(row)
    if blanks:
        issues.append(f"resume text: blank rows {blanks[:10]}")
    if mismatches:
        issues.append(f"resume/raw_text mismatch rows {mismatches[:10]}")


def main() -> None:
    results = [audit_workbook(path) for path in workbook_paths()]
    summary = {
        "workbooks": len(results),
        "passed": sum(result["status"] == "PASS" for result in results),
        "failed": sum(result["status"] == "FAIL" for result in results),
        "results": results,
    }
    print(json.dumps(summary, indent=2))
    raise SystemExit(1 if summary["failed"] else 0)


if __name__ == "__main__":
    main()
