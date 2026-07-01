"""Prepare provenance-safe Excel intake workbooks from the current master pool."""

import argparse
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from backend.excel_intake.workbooks import PROVENANCE_COLUMNS, create_requirement_template  # noqa: E402
MASTER = ROOT / "Reports" / "hirex_candidate_master_with_extracted_resumes.xlsx"
REAL_OUTPUT = ROOT / "storage" / "candidate_pool" / "hirex_candidates.xlsx"
SYNTHETIC_OUTPUT = ROOT / "storage" / "candidate_pool" / "test_synthetic_candidates.xlsx"
JD_OUTPUT = ROOT / "storage" / "job_requirements" / "jd_input.xlsx"


def provenance(row: dict) -> tuple[str, str, str, bool]:
    email = str(row.get("email") or "").lower()
    password_hash = str(row.get("password_hash") or "").lower()
    is_synthetic = email.endswith("@hirexdemo.com") or "$dummy$" in password_hash
    if is_synthetic:
        return "synthetic", "test_only", "HireX synthetic generator", False
    reference = str(row.get("storage_path") or row.get("original_filename") or "")
    return "resume_extracted", "source_extracted_with_synthetic_completion", reference, True


def write_pool(path: Path, headers: list[str], rows: list[list]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidates"
    sheet.append(headers + PROVENANCE_COLUMNS)
    for values in rows:
        sheet.append(values)
    reference = f"A1:{sheet.cell(1, sheet.max_column).column_letter}{max(sheet.max_row, 2)}"
    table = Table(displayName="HireXCandidates", ref=reference)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = reference
    sheet.row_dimensions[1].height = 30
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for index in range(1, sheet.max_column + 1):
        header = sheet.cell(1, index).value
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = 40 if header in {"raw_text", "resume", "source_reference"} else 20
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--include-synthetic",
        action="store_true",
        help="Also generate the optional test-only synthetic candidate workbook.",
    )
    args = parser.parse_args()
    workbook = load_workbook(MASTER, read_only=True, data_only=True)
    sheet = workbook.active
    source_rows = sheet.iter_rows(values_only=True)
    headers = [str(value) for value in next(source_rows)]
    real_rows: list[list] = []
    synthetic_rows: list[list] = []
    for values in source_rows:
        row = dict(zip(headers, values, strict=True))
        extra = list(provenance(row))
        target = synthetic_rows if extra[0] == "synthetic" else real_rows
        target.append(list(values) + extra)
    write_pool(REAL_OUTPUT, headers, real_rows)
    if args.include_synthetic:
        write_pool(SYNTHETIC_OUTPUT, headers, synthetic_rows)
    create_requirement_template(JD_OUTPUT)
    for directory in (ROOT / "storage" / "shortlisted", ROOT / "storage" / "rejected", ROOT / "storage" / "intake_logs"):
        directory.mkdir(parents=True, exist_ok=True)
    print(f"Real candidate rows: {len(real_rows)} -> {REAL_OUTPUT}")
    if args.include_synthetic:
        print(f"Synthetic test rows: {len(synthetic_rows)} -> {SYNTHETIC_OUTPUT}")
    print(f"Job requirement template: {JD_OUTPUT}")


if __name__ == "__main__":
    main()
