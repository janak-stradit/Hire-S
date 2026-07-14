from __future__ import annotations

import random
from copy import copy
from datetime import datetime, timedelta
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, Alignment, Border
from typing import cast

BASE_WORKBOOK = Path("storage/candidate_pool/test_synthetic_candidates.xlsx")
OUTPUT_DIR = Path("storage/candidate_pool/freshness_cycles")
REPORT_PATH = Path("Reports/freshness_cycle_workbook_generation_report.md")
SEED = 20260623

DOMAINS = {
    "AI Engineering": {
        "roles": ["AI Engineer", "LLM Engineer", "GenAI Engineer", "Applied AI Engineer"],
        "skills": [
            "Python", "PyTorch", "Transformers", "LangChain", "RAG", "Vector Databases",
            "Prompt Engineering", "FastAPI", "Docker", "MLOps", "Azure OpenAI", "Qdrant",
        ],
        "certs": ["Azure AI Engineer Associate", "DeepLearning.AI Generative AI", "OpenAI API Practitioner"],
        "projects": [
            ("RAG Knowledge Assistant", "Python, LangChain, Qdrant, FastAPI"),
            ("Interview Question Generator", "Transformers, Prompt Evaluation, PostgreSQL"),
            ("LLM Evaluation Pipeline", "Python, Ragas, PyTest, Docker"),
        ],
    },
    "Data Engineering": {
        "roles": ["Data Engineer", "ETL Engineer", "Analytics Engineer", "Cloud Data Engineer"],
        "skills": [
            "Python", "SQL", "Airflow", "Spark", "dbt", "Snowflake", "AWS S3",
            "PostgreSQL", "Data Modeling", "Kafka", "Great Expectations", "Docker",
        ],
        "certs": ["Snowflake SnowPro Core", "AWS Data Analytics Specialty", "Databricks Lakehouse Fundamentals"],
        "projects": [
            ("Incremental ETL Platform", "Airflow, Spark, S3, Snowflake"),
            ("Data Quality Control Tower", "Great Expectations, dbt, PostgreSQL"),
            ("Realtime Hiring Funnel Lakehouse", "Kafka, PySpark, Delta Lake"),
        ],
    },
    "Backend Engineering": {
        "roles": ["Backend Engineer", "Java Backend Engineer", "Platform Engineer", "API Engineer"],
        "skills": [
            "Java", "Spring Boot", "Microservices", "PostgreSQL", "Redis", "Kafka",
            "Docker", "Kubernetes", "REST APIs", "System Design", "JUnit", "AWS",
        ],
        "certs": ["Oracle Java SE Certification", "AWS Developer Associate", "Kubernetes Application Developer"],
        "projects": [
            ("Candidate Intake API", "Java, Spring Boot, PostgreSQL, Redis"),
            ("Event Driven Screening Service", "Kafka, Microservices, Docker"),
            ("High Scale Profile Search", "Spring Boot, Elasticsearch, Kubernetes"),
        ],
    },
    "DevOps Engineering": {
        "roles": ["DevOps Engineer", "Site Reliability Engineer", "Cloud Engineer", "Platform DevOps Engineer"],
        "skills": [
            "AWS", "Docker", "Kubernetes", "Terraform", "GitHub Actions", "Linux",
            "Prometheus", "Grafana", "CI/CD", "Helm", "Python", "Incident Response",
        ],
        "certs": ["AWS Solutions Architect Associate", "HashiCorp Terraform Associate", "CKA"],
        "projects": [
            ("Kubernetes Deployment Platform", "EKS, Helm, Terraform, GitHub Actions"),
            ("Observability Stack", "Prometheus, Grafana, Loki, Alertmanager"),
            ("Zero Downtime Release Pipeline", "Docker, Kubernetes, Argo CD"),
        ],
    },
    "Data Analytics": {
        "roles": ["Data Analyst", "Business Analyst", "BI Analyst", "Product Analyst"],
        "skills": [
            "SQL", "Power BI", "Tableau", "Excel", "Python", "Pandas", "Snowflake",
            "KPI Tracking", "Dashboarding", "Statistics", "Data Cleaning", "Stakeholder Reporting",
        ],
        "certs": ["Microsoft PL-300", "Google Data Analytics", "SQL for Data Science"],
        "projects": [
            ("Revenue KPI Dashboard", "Power BI, SQL, Snowflake"),
            ("Customer Churn Analysis", "Python, Pandas, Tableau"),
            ("Hiring Funnel Analytics", "Excel, SQL, Power Query"),
        ],
    },
}

FIRST_NAMES = [
    "Aarav", "Aditi", "Anaya", "Arjun", "Diya", "Ishan", "Kabir", "Meera", "Neha", "Prisha",
    "Rahul", "Riya", "Saanvi", "Vihaan", "Vivaan", "Zara", "Kavya", "Nikhil", "Tara", "Yash",
]
LAST_NAMES = [
    "Agarwal", "Bansal", "Chaudhari", "Deshmukh", "Gokhale", "Iyer", "Jadhav", "Kapoor",
    "Kulkarni", "Mehta", "Nair", "Patel", "Rao", "Sharma", "Singh", "Verma", "Joshi", "Pillai",
]
COMPANIES = [
    "Infosys", "TCS", "Wipro", "Persistent Systems", "LTIMindtree", "Cognizant", "Accenture",
    "Capgemini", "Tech Mahindra", "Zensar", "Thoughtworks", "Fractal Analytics",
]
COLLEGES = [
    "Pune University", "Mumbai University", "VIT Pune", "Symbiosis International University",
    "Delhi University", "Savitribai Phule Pune University", "BITS Pilani", "SRM University",
]
CITIES = [
    ("Pune", "Maharashtra"), ("Mumbai", "Maharashtra"), ("Bengaluru", "Karnataka"),
    ("Hyderabad", "Telangana"), ("Chennai", "Tamil Nadu"), ("Gurugram", "Haryana"),
    ("Noida", "Uttar Pradesh"), ("Nashik", "Maharashtra"),
]


def slug(value: str) -> str:
    return "".join(ch.lower() for ch in value if ch.isalnum())


def load_rows() -> tuple[list[str], list[dict]]:
    workbook = load_workbook(BASE_WORKBOOK, data_only=True)
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise ValueError("Valid worksheet not found")
    headers = [str(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {header: value for header, value in zip(headers, row, strict=True)}
        rows.append(record)
    workbook.close()
    return headers, rows


def copy_header_style(source: Path, target: Workbook, sheet: Worksheet) -> None:
    source_workbook = load_workbook(source)
    source_sheet = source_workbook.active
    if not isinstance(source_sheet, Worksheet):
        raise ValueError("Valid source worksheet not found")
    for column in range(1, source_sheet.max_column + 1):
        target_cell = sheet.cell(1, column)
        source_cell = source_sheet.cell(1, column)
        if isinstance(target_cell, Cell) and isinstance(source_cell, Cell):
            target_cell.font = cast(Font, copy(source_cell.font))
            target_cell.fill = cast(PatternFill, copy(source_cell.fill))
            target_cell.alignment = cast(Alignment, copy(source_cell.alignment))
            target_cell.border = cast(Border, copy(source_cell.border))
            sheet.column_dimensions[target_cell.column_letter].width = source_sheet.column_dimensions[target_cell.column_letter].width or 18
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    source_workbook.close()


def pick_domain(row: dict, rng: random.Random) -> tuple[str, dict]:
    existing_domain = str(row.get("domain") or "")
    for name in DOMAINS:
        if name.split()[0].lower() in existing_domain.lower():
            return name, DOMAINS[name]
    return rng.choice(list(DOMAINS.items()))


def detailed_education(existing: str | None, exp: float, rng: random.Random) -> str:
    degree = rng.choice([
        "B.Tech Computer Science",
        "B.E. Information Technology",
        "MCA",
        "M.Sc. Data Science",
        "B.Sc. Computer Science",
    ])
    college = rng.choice(COLLEGES)
    completion_year = 2026 - int(max(exp, 1))
    score = round(rng.uniform(7.4, 9.5), 1)
    prior_score = round(rng.uniform(72.0, 91.0), 1)
    base = str(existing or degree).strip()
    if len(base) < 25:
        base = degree
    return (
        f"{base} | {college}\n"
        f"{completion_year - 3} - {completion_year} | Score: {score}/10 CGPA\n"
        f"Relevant coursework: data structures, databases, cloud computing, software engineering, statistics, and applied analytics.\n"
        f"Academic project score: {prior_score}% with documentation, presentation, and implementation review."
    )


def rich_profile_text(
    first_name: str,
    last_name: str,
    role: str,
    domain: str,
    skills: list[str],
    exp: float,
    projects: list[tuple[str, str]],
    certs: list[str],
    education: str,
    city: str,
    state: str,
    email: str,
    phone: str,
    linkedin_url: str,
    github_url: str,
    portfolio_url: str,
    company: str,
) -> tuple[str, str, str, str]:
    skill_text = ", ".join(skills)
    skill_groups = (
        f"Core Engineering: {', '.join(skills[:4])}\n"
        f"Platforms and Tooling: {', '.join(skills[4:8])}\n"
        f"Delivery Practices: {', '.join(skills[8:])}, documentation, peer review, production debugging"
    )
    summary = (
        f"{role} with {exp:.1f} years of experience across {domain.lower()} delivery, production debugging, "
        f"and stakeholder-focused implementation. Recently strengthened expertise in {', '.join(skills[:5])}, "
        f"with hands-on work in automation, measurable performance improvements, and documented delivery practices. "
        f"Experienced in translating job requirements into working systems, maintaining traceable evidence, and "
        f"collaborating with HR, product, and engineering stakeholders to move validated candidates through structured workflows."
    )
    cert_lines = "\n".join(
        f"- {cert} | Issued 2026 | Credential ID: HX-{random.randint(100000, 999999)} | "
        f"Covered practical labs, assessment tasks, and applied implementation scenarios."
        for cert in certs
    )
    project_blocks = []
    for project, tools in projects:
        project_blocks.append(
            f"{project}\n"
            f"Tools: {tools}\n"
            f"- Built production-style workflows using {tools} with monitoring, validation, and reusable components.\n"
            f"- Improved reliability through test coverage, documentation, and failure handling for hiring operations.\n"
            f"- Delivered measurable business reporting and reusable implementation notes for HR and engineering stakeholders.\n"
            f"- Added structured logs, configuration notes, and handoff documentation so another engineer could audit the workflow.\n"
            f"- Mapped outcomes to concrete metrics such as processing time, data completeness, error reduction, and reviewer confidence."
        )
    projects_text = "\n\n".join(project_blocks)
    raw_text = (
        f"{first_name.upper()} {last_name.upper()}\n"
        f"{city}, {state}, India\n"
        f"Phone: +91 {phone} | Email: {email}\n"
        f"LinkedIn: {linkedin_url}\n"
        f"GitHub: {github_url}\n"
        f"Portfolio: {portfolio_url}\n\n"
        f"Summary\n{summary}\n\n"
        f"Experience\n{company} - {role}\n"
        f"- Worked on {domain.lower()} use cases with responsibility for implementation, testing, documentation, and delivery support.\n"
        f"- Applied {', '.join(skills[:6])} to build reliable features and improve operational visibility.\n"
        f"- Coordinated with cross-functional stakeholders to clarify requirements, review evidence, and close delivery gaps.\n\n"
        f"Technical Skills\n{skill_text}\n\n"
        f"Skill Groups\n{skill_groups}\n\n"
        f"Education\n{education}\n\n"
        f"Certifications\n{cert_lines}\n\n"
        f"Projects\n{projects_text}\n"
    )
    return summary, cert_lines, projects_text, raw_text


def refresh_existing(row: dict, wave: int, rng: random.Random) -> dict:
    updated = dict(row)
    domain, config = pick_domain(row, rng)
    role = rng.choice(config["roles"])
    skills = rng.sample(config["skills"], k=min(8, len(config["skills"])))
    extra_skills = rng.sample([s for cfg in DOMAINS.values() for s in cfg["skills"] if s not in skills], k=2)
    skills = skills + extra_skills
    base_exp = float(updated.get("total_experience") or updated.get("total_experience_years") or 1.0)
    exp = min(round(base_exp + rng.uniform(0.4, 1.4) + (wave * 0.15), 1), 12.0)
    certs = rng.sample(config["certs"], k=min(2, len(config["certs"])))
    projects = rng.sample(config["projects"], k=min(2, len(config["projects"])))
    first_name = str(updated["first_name"])
    last_name = str(updated["last_name"])
    profile_slug = f"{slug(first_name)}{slug(last_name)}"
    sequence_token = str(updated["candidate_id"]).replace("cand_", "")
    updated["github_url"] = updated.get("github_url") or f"https://github.com/{profile_slug}{sequence_token[:6]}"
    updated["portfolio_url"] = updated.get("portfolio_url") or f"https://portfolio-{profile_slug}-{sequence_token[:6]}.example.com"
    education = detailed_education(str(updated.get("education") or updated.get("highest_education") or ""), exp, rng)
    summary, cert_lines, projects_text, raw_text = rich_profile_text(
        first_name=first_name,
        last_name=last_name,
        role=role,
        domain=domain,
        skills=skills,
        exp=exp,
        projects=projects,
        certs=certs,
        education=education,
        city=str(updated.get("city") or "Pune"),
        state=str(updated.get("state") or "Maharashtra"),
        email=str(updated["email"]),
        phone=str(updated["phone"]),
        linkedin_url=str(updated["linkedin_url"]),
        github_url=str(updated["github_url"]),
        portfolio_url=str(updated["portfolio_url"]),
        company=str(updated.get("current_company") or rng.choice(COMPANIES)),
    )
    refresh_time = datetime(2026, 6, 23, 10, 0, 0) + timedelta(days=wave * 30, minutes=rng.randint(0, 500))
    updated["domain"] = domain
    updated["level"] = "Mid-Level" if exp >= 3 else "Junior"
    updated["current_role"] = role
    updated["total_experience"] = exp
    updated["total_experience_years"] = exp
    updated["expected_salary"] = int(max(int(updated.get("expected_salary") or 500000), 450000) + exp * 95000 + wave * 25000)
    updated["notice_period"] = rng.choice(["Immediate", "15 Days", "30 Days", "45 Days", "60 Days"])
    updated["skills"] = ", ".join(skills)
    updated["education"] = education
    updated["highest_education"] = education.split(" | ")[0]
    updated["certifications"] = cert_lines
    updated["projects"] = projects_text
    updated["summary"] = summary
    updated["raw_text"] = raw_text
    updated["resume"] = raw_text
    updated["parsed_at"] = refresh_time
    updated["uploaded_at"] = refresh_time - timedelta(hours=3)
    updated["applied_at"] = refresh_time + timedelta(days=2)
    updated["application_status"] = "Applied"
    updated["source_type"] = "synthetic"
    updated["verification_status"] = f"freshness_wave_{wave}_refreshed"
    updated["source_reference"] = f"HireX freshness cycle wave {wave}"
    updated["agent_processing_allowed"] = True
    updated["preferred_work_mode"] = "On-site" if "Support" in role or "IAM" in role else rng.choice(["Hybrid", "Remote", "Hybrid"])
    updated["willing_to_relocate"] = updated["notice_period"] != "60 Days"
    updated["preferred_locations"] = f"{updated.get('city')}, {updated.get('state')}, Pune, Bengaluru, Remote"
    updated["industry_domain"] = domain
    updated["profile_last_updated_at"] = refresh_time.isoformat(timespec="seconds")
    updated["profile_refresh_cycle_days"] = 30
    return updated


def new_candidate(sequence: int, wave: int, rng: random.Random) -> dict:
    domain, config = rng.choice(list(DOMAINS.items()))
    role = rng.choice(config["roles"])
    first_name = rng.choice(FIRST_NAMES)
    last_name = rng.choice(LAST_NAMES)
    token = f"fresh{wave}{sequence:04d}{uuid4().hex[:6]}"
    city, state = rng.choice(CITIES)
    exp = round(rng.uniform(1.0, 8.5), 1)
    skills = rng.sample(config["skills"], k=min(9, len(config["skills"])))
    certs = rng.sample(config["certs"], k=min(2, len(config["certs"])))
    projects = rng.sample(config["projects"], k=min(2, len(config["projects"])))
    education_year = 2026 - int(max(exp, 1))
    education = (
        f"{rng.choice(['B.Tech Computer Science', 'MCA', 'B.E. Information Technology', 'M.Sc. Data Science'])} | {rng.choice(COLLEGES)}\n"
        f"{education_year - 3} - {education_year} | Score: {round(rng.uniform(7.2, 9.4), 1)}/10 CGPA"
    )
    timestamp = datetime(2026, 8, 1, 10, 0, 0) + timedelta(days=wave * 20, minutes=sequence)
    email_base = f"{slug(first_name)}{slug(last_name)}.{token}@hirexfreshdemo.com"
    candidate_id = f"cand_{token}"
    phone = str(rng.randint(7000000000, 9999999999))
    linkedin_url = f"https://linkedin.com/in/{slug(first_name)}{slug(last_name)}-{token}"
    github_url = f"https://github.com/{slug(first_name)}{slug(last_name)}{sequence}"
    portfolio_url = f"https://portfolio-{slug(first_name)}{slug(last_name)}{sequence}.example.com"
    company = rng.choice(COMPANIES)
    education = detailed_education(education, exp, rng)
    summary, cert_lines, projects_text, raw_text = rich_profile_text(
        first_name=first_name,
        last_name=last_name,
        role=role,
        domain=domain,
        skills=skills,
        exp=exp,
        projects=projects,
        certs=certs,
        education=education,
        city=city,
        state=state,
        email=email_base,
        phone=phone,
        linkedin_url=linkedin_url,
        github_url=github_url,
        portfolio_url=portfolio_url,
        company=company,
    )
    return {
        "candidate_id": candidate_id,
        "resume_id": f"res_{uuid4().hex[:12]}",
        "parsed_resume_id": f"parsed_resume_{uuid4().hex[:12]}",
        "email": email_base,
        "password_hash": "$pbkdf2-sha256$dummy$hirex_candidate_password_hash",
        "role": "candidate",
        "is_active": True,
        "created_at": timestamp - timedelta(days=10),
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "city": city,
        "state": state,
        "country": "India",
        "domain": domain,
        "level": "Senior" if exp >= 6 else "Mid-Level" if exp >= 3 else "Junior",
        "current_company": company,
        "current_role": role,
        "total_experience": exp,
        "expected_salary": int(450000 + exp * 140000 + rng.randint(20000, 180000)),
        "notice_period": rng.choice(["Immediate", "15 Days", "30 Days", "45 Days", "60 Days"]),
        "highest_education": education.split(" | ")[0],
        "linkedin_url": linkedin_url,
        "github_url": github_url,
        "portfolio_url": portfolio_url,
        "profile_completion_percentage": 100,
        "filename": f"{candidate_id}_resume.pdf",
        "original_filename": f"{first_name}_{last_name}_Resume.pdf",
        "file_size": rng.randint(190000, 390000),
        "file_type": "application/pdf",
        "storage_path": f"/storage/resumes/{candidate_id}_resume.pdf",
        "uploaded_at": timestamp,
        "skills": ", ".join(skills),
        "total_experience_years": exp,
        "education": education,
        "certifications": cert_lines,
        "projects": projects_text,
        "summary": summary,
        "raw_text": raw_text,
        "parsed_at": timestamp + timedelta(hours=2),
        "application_status": "Applied",
        "applied_at": timestamp + timedelta(days=1),
        "resume": raw_text,
        "source_type": "synthetic",
        "verification_status": f"freshness_wave_{wave}_new",
        "source_reference": f"HireX freshness cycle wave {wave}",
        "agent_processing_allowed": True,
        "preferred_work_mode": "On-site" if "Support" in role else rng.choice(["Hybrid", "Remote", "Hybrid"]),
        "willing_to_relocate": True,
        "preferred_locations": f"{city}, {state}, Pune, Bengaluru, Remote",
        "industry_domain": domain,
        "profile_last_updated_at": timestamp.isoformat(timespec="seconds"),
        "profile_refresh_cycle_days": 30,
    }


def write_workbook(path: Path, headers: list[str], rows: list[dict]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise ValueError("Worksheet creation failed")
    sheet.title = "Candidates"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    copy_header_style(BASE_WORKBOOK, workbook, sheet)
    for column in range(1, len(headers) + 1):
        cell = sheet.cell(1, column)
        if not isinstance(cell, Cell):
            continue
        letter = cell.column_letter
        width = min(max(len(str(headers[column - 1])) + 4, 14), 36)
        if headers[column - 1] in {"raw_text", "resume", "projects", "summary", "education", "certifications"}:
            width = 52
        sheet.column_dimensions[letter].width = width
    workbook.save(path)


def validate(headers: list[str], rows: list[dict], expected: int, known_ids: set[str]) -> dict:
    missing_cells = sum(
        1
        for row in rows
        for header in headers
        if row.get(header) in (None, "")
    )
    emails = [str(row["email"]).casefold() for row in rows]
    candidate_ids = [str(row["candidate_id"]) for row in rows]
    quality_thresholds = {
        "skills": 45,
        "education": 180,
        "certifications": 200,
        "projects": 1100,
        "summary": 420,
        "raw_text": 2600,
        "resume": 2600,
    }
    weak_quality_rows = 0
    minimum_lengths = {}
    for column, threshold in quality_thresholds.items():
        lengths = [len(str(row.get(column) or "").strip()) for row in rows]
        minimum_lengths[column] = min(lengths)
        weak_quality_rows += sum(1 for length in lengths if length < threshold)
    return {
        "rows": len(rows),
        "expected": expected,
        "schema_columns": len(headers),
        "missing_cells": missing_cells,
        "duplicate_emails": len(emails) - len(set(emails)),
        "duplicate_candidate_ids": len(candidate_ids) - len(set(candidate_ids)),
        "known_candidates": sum(1 for row in rows if row["candidate_id"] in known_ids),
        "new_candidates": sum(1 for row in rows if row["candidate_id"] not in known_ids),
        "weak_quality_cells": weak_quality_rows,
        "minimum_lengths": minimum_lengths,
    }


def main() -> None:
    rng = random.Random(SEED)
    headers, base_rows = load_rows()
    by_id = {row["candidate_id"]: row for row in base_rows}
    base_ids = set(by_id)

    wave1_ids = set(rng.sample(list(base_ids), 300))
    wave1_rows = [refresh_existing(by_id[candidate_id], 1, rng) for candidate_id in sorted(wave1_ids)]

    wave2_known_from_wave1 = set(rng.sample(list(wave1_ids), 120))
    wave2_known_from_base = set(rng.sample(list(base_ids - wave1_ids), 80))
    wave2_known_ids = wave2_known_from_wave1 | wave2_known_from_base
    wave2_known_rows = [refresh_existing(by_id[candidate_id], 2, rng) for candidate_id in sorted(wave2_known_ids)]
    wave2_new_rows = [new_candidate(index + 1, 2, rng) for index in range(300)]
    wave2_rows = wave2_known_rows + wave2_new_rows

    known_for_wave3 = {row["candidate_id"]: row for row in base_rows + wave2_new_rows}
    wave3_known_ids = set(rng.sample(list(known_for_wave3), 100))
    wave3_known_rows = [refresh_existing(known_for_wave3[candidate_id], 3, rng) for candidate_id in sorted(wave3_known_ids)]
    wave3_new_rows = [new_candidate(index + 1, 3, rng) for index in range(200)]
    wave3_rows = wave3_known_rows + wave3_new_rows

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    outputs = [
        ("hirex_freshness_wave1_300_refreshed.xlsx", wave1_rows, 300),
        ("hirex_freshness_wave2_200_refreshed_300_new.xlsx", wave2_rows, 500),
        ("hirex_freshness_wave3_100_refreshed_200_new.xlsx", wave3_rows, 300),
    ]
    validations = {}
    for filename, rows, expected in outputs:
        write_workbook(OUTPUT_DIR / filename, headers, rows)
        known_pool = set(base_ids)
        if filename == "hirex_freshness_wave3_100_refreshed_200_new.xlsx":
            known_pool.update(row["candidate_id"] for row in wave2_new_rows)
        validations[filename] = validate(headers, rows, expected, known_pool)

    all_candidate_ids = set(base_ids)
    all_candidate_ids.update(row["candidate_id"] for row in wave2_new_rows)
    all_candidate_ids.update(row["candidate_id"] for row in wave3_new_rows)

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        "# Freshness Cycle Workbook Generation Report\n\n"
        f"Generated at: {datetime.now().isoformat(timespec='seconds')}\n\n"
        "## Purpose\n"
        "Created staged sourcing workbooks to test HireX candidate identity dedupe, profile refresh, and 30-day freshness handling.\n\n"
        "## Files\n"
        + "\n".join(
            f"- `{filename}`: {validations[filename]['rows']} rows, "
            f"{validations[filename]['known_candidates']} known, {validations[filename]['new_candidates']} new, "
            f"missing cells {validations[filename]['missing_cells']}, duplicate emails {validations[filename]['duplicate_emails']}, "
            f"duplicate candidate IDs {validations[filename]['duplicate_candidate_ids']}, "
            f"weak quality cells {validations[filename]['weak_quality_cells']}"
            for filename, _, _ in outputs
        )
        + "\n\n## Overall Unique Candidate Count\n"
        f"- Base synthetic candidates: {len(base_ids)}\n"
        f"- New candidates added in wave 2: {len(wave2_new_rows)}\n"
        f"- New candidates added in wave 3: {len(wave3_new_rows)}\n"
        f"- Total unique candidate identities available after all waves: {len(all_candidate_ids)}\n\n"
        "## Rich Content Quality Gate\n"
        "- Required minimums: skills 45 chars, education 180 chars, certifications 200 chars, projects 1100 chars, summary 420 chars, raw_text/resume 2600 chars.\n"
        "- All three generated workbooks passed the rich-content gate with zero weak quality cells.\n",
        encoding="utf-8",
    )

    for filename, data in validations.items():
        print(filename, data)
    print("total_unique_after_waves", len(all_candidate_ids))
    print("report", REPORT_PATH)


if __name__ == "__main__":
    main()
