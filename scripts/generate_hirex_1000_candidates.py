import math
import random
import re
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from faker import Faker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


fake = Faker("en_IN")
Faker.seed(42)
random.seed(42)

TOTAL_CANDIDATES = 1000
BASE_DATE = datetime(2026, 6, 19, 10, 0, 0)
OUTPUT_FILE = Path(__file__).resolve().parents[1] / "Reports" / "hirex_1000_candidate_dummy_data.xlsx"


DOMAIN_BLUEPRINTS = {
    "IT Support": {
        "roles": ["IT Support Engineer", "System Administrator", "Network Engineer", "Desktop Support Engineer", "IT Manager"],
        "skills": {
            "Systems": ["Windows Server", "Linux", "Active Directory", "Microsoft 365"],
            "Networking": ["TCP/IP", "DNS", "DHCP", "VPN", "LAN/WAN"],
            "Operations": ["ServiceNow", "Incident Management", "Asset Management", "Troubleshooting"],
        },
        "projects": [
            ("Enterprise Service Desk Optimization", "ServiceNow, PowerShell, Active Directory", "standardized ticket triage and automated repetitive account-support activities"),
            ("Secure Office Network Rollout", "Cisco, DNS, DHCP, VPN", "designed segmented office connectivity and strengthened remote-access reliability"),
            ("IT Asset Lifecycle Dashboard", "Excel, Power BI, SQL", "created a centralized view of device ownership, warranty, and compliance status"),
        ],
        "certifications": ["CompTIA A+", "Cisco CCNA", "Microsoft Azure Fundamentals AZ-900"],
        "outcome": "reliable end-user services, secure infrastructure, and faster incident resolution",
        "bullets": [
            "Resolved {volume}+ monthly L1/L2 incidents across Windows, Linux, network, and access-management queues while maintaining {metric}% SLA compliance.",
            "Automated account provisioning and routine health checks using {skill}, reducing manual support effort by {gain}%.",
            "Diagnosed recurring DNS, DHCP, VPN, and endpoint issues and lowered repeat incidents by {gain}% through root-cause documentation.",
            "Managed hardware and software inventory for {volume}+ assets and improved audit accuracy to {metric}%.",
            "Published knowledge-base articles and trained {team}+ users on secure access, collaboration tools, and self-service troubleshooting.",
        ],
    },
    "Software Development": {
        "roles": ["Python Developer", "Java Developer", "Backend Developer", "Full Stack Developer", "Engineering Manager"],
        "skills": {
            "Languages": ["Python", "Java", "TypeScript", "SQL"],
            "Frameworks": ["FastAPI", "Django", "React", "Spring Boot"],
            "Engineering": ["PostgreSQL", "REST API", "Git", "Docker", "PyTest"],
        },
        "projects": [
            ("Talent Application Platform", "Python, FastAPI, PostgreSQL, React", "built secure candidate, job, resume, and application workflows with role-based access"),
            ("Inventory Management API", "Java, Spring Boot, PostgreSQL, Docker", "implemented transactional inventory services and auditable stock movement APIs"),
            ("Identity and Access Service", "FastAPI, JWT, Redis, PyTest", "delivered token-based authentication, authorization, and session controls"),
        ],
        "certifications": ["Python Institute PCEP", "AWS Certified Cloud Practitioner", "Oracle Java SE Certification"],
        "outcome": "maintainable software, secure APIs, and dependable product delivery",
        "bullets": [
            "Designed and delivered {volume}+ REST endpoints using {skill}, improving feature throughput by {gain}% while preserving API contracts.",
            "Optimized PostgreSQL queries and indexes, reducing p95 response time by {gain}% across high-traffic workflows.",
            "Implemented automated unit and integration tests with {skill}, raising critical-path coverage to {metric}%.",
            "Containerized services and standardized CI checks, reducing deployment preparation from {hours} hours to under {reduced} hours.",
            "Partnered with product, QA, and design teams in {team}-member delivery squads to translate requirements into production releases.",
        ],
    },
    "Data Analytics": {
        "roles": ["Data Analyst", "BI Analyst", "Power BI Developer", "Analytics Consultant", "Analytics Manager"],
        "skills": {
            "Analytics": ["SQL", "Python", "Pandas", "NumPy", "Data Cleaning"],
            "Visualization": ["Power BI", "Tableau", "Excel", "KPI Tracking"],
            "Data Platforms": ["Snowflake", "PostgreSQL", "Salesforce", "AWS S3"],
        },
        "projects": [
            ("Customer Churn Intelligence", "SQL, Python, Pandas, Tableau", "identified high-risk segments and translated churn drivers into retention actions"),
            ("Air Quality Analytics Platform", "Snowflake, Snowpark, Streamlit, AWS S3", "combined semi-structured environmental data for city and pollutant trend analysis"),
            ("Executive Revenue Dashboard", "Power BI, SQL, Excel, DAX", "delivered governed KPI reporting with drilldowns for revenue, margin, and regional performance"),
        ],
        "certifications": ["Microsoft Power BI Data Analyst PL-300", "Snowflake SnowPro Core", "Google Data Analytics Professional Certificate"],
        "outcome": "trusted metrics, decision-ready dashboards, and measurable business insights",
        "bullets": [
            "Integrated {volume}+ source tables from Snowflake, Salesforce, SharePoint, and S3 into validated analytics datasets using {skill}.",
            "Built Power BI and Tableau dashboards for {team}+ stakeholders, improving weekly KPI review time by {gain}%.",
            "Automated data-quality and reconciliation checks with {skill}, increasing reporting accuracy to {metric}%.",
            "Optimized incremental transformations and reduced refresh duration from {hours} hours to {reduced} hours.",
            "Presented trend, variance, and cohort findings that supported planning decisions worth INR {value} lakh.",
        ],
    },
    "Data Engineering": {
        "roles": ["Data Engineer", "ETL Developer", "Big Data Engineer", "Data Platform Engineer", "Data Engineering Manager"],
        "skills": {
            "Engineering": ["Python", "SQL", "PySpark", "ETL", "Data Modeling"],
            "Platforms": ["Apache Airflow", "Databricks", "Azure Data Factory", "Snowflake"],
            "Storage": ["PostgreSQL", "AWS S3", "Delta Lake", "Kafka"],
        },
        "projects": [
            ("Enterprise Lakehouse Pipeline", "PySpark, Databricks, Delta Lake, Airflow", "implemented medallion pipelines with schema enforcement, lineage, and incremental processing"),
            ("Cloud Data Warehouse Migration", "Snowflake, Python, SQL, dbt", "migrated legacy reporting data into tested dimensional models"),
            ("Near Real-Time Event Platform", "Kafka, Python, PostgreSQL, AWS", "processed operational events with resilient ingestion and observable delivery"),
        ],
        "certifications": ["Microsoft Azure Data Engineer DP-203", "Databricks Lakehouse Fundamentals", "AWS Certified Data Engineer Associate"],
        "outcome": "reliable data products, scalable pipelines, and governed analytics platforms",
        "bullets": [
            "Built and operated {volume}+ batch and streaming pipelines using {skill}, processing {records}+ million records per day.",
            "Introduced incremental loading and partition strategies that reduced end-to-end processing time by {gain}%.",
            "Developed reconciliation, schema, and freshness checks with {skill}, maintaining {metric}% successful daily runs.",
            "Modeled analytics-ready fact and dimension tables for {team}+ business reporting use cases.",
            "Implemented monitoring and recovery playbooks that reduced mean time to restore failed pipelines by {gain}%.",
        ],
    },
    "Testing": {
        "roles": ["QA Tester", "Manual Tester", "Automation Tester", "QA Lead", "Test Manager"],
        "skills": {
            "Testing": ["Manual Testing", "Regression Testing", "API Testing", "Exploratory Testing"],
            "Automation": ["Selenium", "PyTest", "Postman", "REST Assured"],
            "Delivery": ["JIRA", "TestRail", "Git", "CI/CD"],
        },
        "projects": [
            ("E-commerce Quality Suite", "Selenium, PyTest, Postman, JIRA", "automated checkout, payment, catalog, and account regression coverage"),
            ("API Contract Verification", "Postman, REST Assured, SQL", "validated service contracts, authorization, errors, and database side effects"),
            ("Release Readiness Framework", "TestRail, JIRA, CI/CD", "standardized risk-based test planning and release evidence"),
        ],
        "certifications": ["ISTQB Certified Tester Foundation Level", "Selenium WebDriver Certification", "Postman API Fundamentals Student Expert"],
        "outcome": "predictable releases, strong defect prevention, and measurable product quality",
        "bullets": [
            "Authored and maintained {volume}+ functional, integration, regression, and negative test cases across web and API workflows.",
            "Automated critical user journeys using {skill}, reducing manual regression time by {gain}%.",
            "Validated REST APIs, authentication, and database changes and maintained {metric}% pass reliability in CI runs.",
            "Triaged defects with engineering and product teams, reducing defect leakage by {gain}% across quarterly releases.",
            "Led release-readiness reviews for {team}+ stakeholders with traceable coverage, risks, and evidence in JIRA and TestRail.",
        ],
    },
    "DevOps": {
        "roles": ["DevOps Engineer", "Cloud Engineer", "Site Reliability Engineer", "DevOps Lead", "DevOps Manager"],
        "skills": {
            "Cloud": ["AWS", "Azure", "Kubernetes", "Docker"],
            "Automation": ["Terraform", "Jenkins", "GitHub Actions", "Ansible"],
            "Reliability": ["Linux", "Prometheus", "Grafana", "Incident Response"],
        },
        "projects": [
            ("Container Platform Modernization", "Kubernetes, Docker, Helm, Prometheus", "standardized service deployment, scaling, observability, and rollback controls"),
            ("Cloud Infrastructure as Code", "Terraform, AWS, GitHub Actions", "provisioned repeatable network, compute, storage, and security environments"),
            ("Deployment Reliability Program", "Jenkins, Argo CD, Grafana", "introduced progressive delivery, quality gates, and release monitoring"),
        ],
        "certifications": ["AWS Solutions Architect Associate", "Certified Kubernetes Administrator", "HashiCorp Terraform Associate"],
        "outcome": "repeatable delivery, resilient cloud operations, and observable services",
        "bullets": [
            "Managed {volume}+ cloud workloads using {skill} with documented availability, backup, and recovery controls.",
            "Automated infrastructure provisioning with {skill}, reducing environment setup time by {gain}%.",
            "Built CI/CD pipelines with security and quality gates, increasing deployment frequency by {gain}%.",
            "Implemented logs, metrics, alerts, and runbooks that improved service availability to {metric}%.",
            "Coordinated incident response across {team}+ service owners and reduced mean time to recovery from {hours} hours to {reduced} hours.",
        ],
    },
    "Finance": {
        "roles": ["Finance Associate", "Financial Analyst", "Investment Analyst", "Finance Manager", "VP Finance"],
        "skills": {
            "Planning": ["Financial Modeling", "Budgeting", "Forecasting", "Variance Analysis"],
            "Analysis": ["Excel", "Power BI", "Risk Analysis", "Valuation"],
            "Operations": ["SAP", "Tally", "Management Reporting", "Internal Controls"],
        },
        "projects": [
            ("Integrated Budget and Forecast Model", "Excel, Power Query, Power BI", "connected revenue, cost, headcount, cash flow, and scenario assumptions"),
            ("Investment Risk Review", "Excel, SQL, Power BI", "evaluated return, concentration, liquidity, and downside exposure"),
            ("Working Capital Optimization", "SAP, Excel, Power BI", "analyzed receivables, payables, inventory, and cash-conversion drivers"),
        ],
        "certifications": ["CFA Level I", "NISM Research Analyst", "Financial Modeling and Valuation Analyst"],
        "outcome": "accurate planning, controlled financial operations, and decision-ready analysis",
        "bullets": [
            "Prepared monthly forecasts and variance analysis for an INR {value} crore portfolio using {skill}.",
            "Built driver-based financial models that improved forecast accuracy to {metric}%.",
            "Automated management reporting and reduced monthly close-analysis effort by {gain}%.",
            "Evaluated investment, liquidity, and operating risks and presented recommendations to {team}+ senior stakeholders.",
            "Strengthened reconciliations and controls across {volume}+ cost centers with documented exception resolution.",
        ],
    },
    "HR": {
        "roles": ["HR Executive", "Recruiter", "Talent Acquisition Specialist", "HR Manager", "VP HR"],
        "skills": {
            "Talent": ["Recruitment", "Candidate Screening", "Interview Coordination", "Onboarding"],
            "People Operations": ["Payroll", "Employee Engagement", "HRMS", "Policy Administration"],
            "Analytics": ["HR Analytics", "Excel", "Power BI", "Workforce Reporting"],
        },
        "projects": [
            ("Recruitment Operations Tracker", "Excel, Power BI, ATS", "tracked pipeline conversion, aging, source quality, and recruiter productivity"),
            ("Digital Employee Onboarding", "HRMS, Forms, Power Automate", "standardized documentation, approvals, induction, and probation checkpoints"),
            ("Engagement and Attrition Analytics", "Excel, SQL, Power BI", "identified workforce trends and prioritized manager actions"),
        ],
        "certifications": ["SHRM Certified Professional", "LinkedIn Talent Solutions Recruiter", "People Analytics Certification"],
        "outcome": "efficient hiring, consistent people operations, and evidence-based workforce decisions",
        "bullets": [
            "Managed {volume}+ open positions across sourcing, screening, coordination, offer, and joining workflows.",
            "Improved interview-to-offer conversion by {gain}% through structured screening and stakeholder calibration.",
            "Reduced onboarding turnaround by {gain}% using {skill} workflows and standardized document controls.",
            "Produced workforce, attrition, diversity, and hiring reports with {metric}% data completeness.",
            "Partnered with {team}+ managers on engagement, performance, policy, and employee-relations actions.",
        ],
    },
    "Accounting": {
        "roles": ["Account Assistant", "Accountant", "Senior Accountant", "Accounts Manager", "Finance Controller"],
        "skills": {
            "Accounting": ["General Ledger", "Accounts Payable", "Accounts Receivable", "Balance Sheet"],
            "Compliance": ["GST", "Taxation", "Auditing", "TDS"],
            "Systems": ["Tally", "Excel", "SAP", "Bank Reconciliation"],
        },
        "projects": [
            ("GST Compliance Automation", "Tally, Excel, Power Query", "reconciled invoices, tax ledgers, returns, and exception follow-up"),
            ("Month-End Close Tracker", "Excel, SAP, Power BI", "controlled journal, reconciliation, accrual, and review milestones"),
            ("Vendor Payment Governance", "Tally, Excel, Banking Portal", "tracked approvals, due dates, deductions, and payment evidence"),
        ],
        "certifications": ["Tally Prime Certification", "GST Practitioner Certificate", "Diploma in IFRS"],
        "outcome": "accurate books, timely compliance, and controlled financial reporting",
        "bullets": [
            "Processed and reconciled {volume}+ monthly transactions across receivables, payables, bank, and general-ledger accounts.",
            "Prepared GST, TDS, and statutory schedules with {metric}% on-time filing compliance.",
            "Reduced reconciliation exceptions by {gain}% through standardized templates and {skill} controls.",
            "Supported month-end close for an INR {value} crore business and delivered schedules within {reduced} working days.",
            "Coordinated audit evidence and resolved {volume}+ ledger queries with vendors, operations, and finance reviewers.",
        ],
    },
    "Sales and Marketing": {
        "roles": ["Sales Executive", "Marketing Executive", "Business Development Associate", "Sales Manager", "VP Sales"],
        "skills": {
            "Growth": ["Lead Generation", "Digital Marketing", "SEO", "Google Ads"],
            "Sales": ["CRM", "Cold Calling", "Negotiation", "Pipeline Management"],
            "Analytics": ["Google Analytics", "Excel", "Power BI", "Campaign Reporting"],
        },
        "projects": [
            ("Integrated Lead Generation Campaign", "Google Ads, SEO, HubSpot, Analytics", "connected audience targeting, content, landing pages, and lead nurturing"),
            ("CRM Pipeline Transformation", "Salesforce, Excel, Power BI", "standardized opportunity stages, activity tracking, and forecast governance"),
            ("Regional Market Expansion", "Market Research, CRM, Power BI", "prioritized accounts, channels, territories, and launch activities"),
        ],
        "certifications": ["Google Digital Marketing and E-commerce", "HubSpot Sales Software", "Google Ads Search Certification"],
        "outcome": "qualified pipeline growth, stronger conversion, and measurable campaign performance",
        "bullets": [
            "Generated INR {value} lakh in qualified pipeline through {skill}, outbound prospecting, and partner channels.",
            "Managed {volume}+ active opportunities in CRM and improved stage-to-stage conversion by {gain}%.",
            "Optimized digital campaigns using audience, keyword, creative, and landing-page analysis, reducing cost per lead by {gain}%.",
            "Built weekly funnel and forecast reporting for {team}+ sales and marketing stakeholders.",
            "Negotiated and closed strategic accounts while maintaining {metric}% CRM activity and forecast hygiene.",
        ],
    },
}


LEVELS = [
    ("Fresher", 0.2, 1.0),
    ("Associate", 1.0, 3.0),
    ("Executive", 2.0, 5.0),
    ("Senior", 4.0, 8.0),
    ("Lead", 6.0, 10.0),
    ("Manager", 8.0, 14.0),
    ("Senior Manager", 10.0, 16.0),
    ("Director", 12.0, 20.0),
    ("VP", 15.0, 25.0),
]

EDUCATION_BY_DOMAIN = {
    "IT Support": ["B.Sc. Computer Science", "BCA", "Diploma in Information Technology", "B.E. Computer Engineering"],
    "Software Development": ["BCA", "MCA", "B.E. Computer Engineering", "B.Tech. Information Technology"],
    "Data Analytics": ["B.Sc. Statistics", "M.Sc. Data Science", "B.Tech. Information Technology", "MBA Business Analytics"],
    "Data Engineering": ["MCA", "B.E. Computer Engineering", "B.Tech. Information Technology", "M.Sc. Data Science"],
    "Testing": ["BCA", "MCA", "B.E. Computer Engineering", "B.Sc. Information Technology"],
    "DevOps": ["B.E. Computer Engineering", "B.Tech. Information Technology", "MCA", "B.Sc. Computer Science"],
    "Finance": ["B.Com.", "M.Com.", "MBA Finance", "BBA Finance"],
    "HR": ["MBA Human Resources", "BBA Human Resources", "M.Com.", "B.A. Psychology"],
    "Accounting": ["B.Com.", "M.Com.", "MBA Finance", "CA Intermediate"],
    "Sales and Marketing": ["BBA Marketing", "MBA Marketing", "B.Com.", "B.A. Economics"],
}

COMPANIES = [
    "Infosys", "TCS", "Wipro", "Cognizant", "Accenture", "Capgemini", "Tech Mahindra",
    "Persistent Systems", "LTIMindtree", "HCLTech", "Deloitte", "KPMG", "EY", "PwC",
    "ICICI Bank", "HDFC Bank", "Kotak Mahindra Bank", "Bajaj Finance", "Zomato", "Paytm",
    "PhonePe", "Flipkart", "Amazon", "Freshworks", "Zoho", "PubMatic", "Cybage", "Druva", "Icertis",
]

CITIES_STATES = [
    ("Pune", "Maharashtra"), ("Mumbai", "Maharashtra"), ("Bengaluru", "Karnataka"),
    ("Hyderabad", "Telangana"), ("Chennai", "Tamil Nadu"), ("Delhi", "Delhi"),
    ("Noida", "Uttar Pradesh"), ("Gurugram", "Haryana"), ("Ahmedabad", "Gujarat"),
    ("Indore", "Madhya Pradesh"), ("Nagpur", "Maharashtra"), ("Nashik", "Maharashtra"),
]

INSTITUTIONS = [
    "Savitribai Phule Pune University", "University of Mumbai", "Anna University",
    "Osmania University", "Visvesvaraya Technological University", "Delhi University",
    "Symbiosis International University", "Dr. A.P.J. Abdul Kalam Technical University",
]

NOTICE_PERIODS = ["Immediate", "15 Days", "30 Days", "45 Days", "60 Days", "90 Days"]
APPLICATION_STATUSES = [
    "Applied", "Under Review", "Shortlisted", "Rejected", "R1 Scheduled",
    "R1 Completed", "T1 Scheduled", "T1 Completed", "Final Review",
]

FINAL_COLUMNS = [
    "candidate_id", "resume_id", "parsed_resume_id", "email", "password_hash", "role",
    "is_active", "created_at", "first_name", "last_name", "phone", "city", "state",
    "country", "domain", "level", "current_company", "current_role", "total_experience",
    "expected_salary", "notice_period", "highest_education", "linkedin_url", "github_url",
    "portfolio_url", "profile_completion_percentage", "filename", "original_filename",
    "file_size", "file_type", "storage_path", "uploaded_at", "skills",
    "total_experience_years", "education", "certifications", "projects", "summary",
    "raw_text", "parsed_at", "application_status", "applied_at", "resume",
]


def deterministic_id(prefix: str, index: int) -> str:
    value = uuid.uuid5(uuid.NAMESPACE_URL, f"hirex-demo-{prefix}-{index}").hex[:12]
    return f"{prefix}_{value}"


def salary_by_experience(experience: float) -> int:
    bands = [(1, 250000, 500000), (3, 400000, 800000), (6, 700000, 1500000),
             (10, 1200000, 2500000), (15, 2200000, 4000000), (100, 3500000, 7500000)]
    for ceiling, low, high in bands:
        if experience <= ceiling:
            return random.randrange(low, high + 1, 25000)
    return 7500000


def flatten_skills(skill_groups: dict[str, list[str]]) -> list[str]:
    return [skill for group in skill_groups.values() for skill in group]


def format_month(value: datetime) -> str:
    return value.strftime("%b %Y")


def shift_months(value: datetime, months: int) -> datetime:
    month_index = value.year * 12 + value.month - 1 + months
    year, month_zero = divmod(month_index, 12)
    return value.replace(year=year, month=month_zero + 1, day=1)


def make_impact_bullet(template: str, selected_skills: list[str]) -> str:
    return template.format(
        skill=random.choice(selected_skills),
        volume=random.choice([25, 40, 60, 80, 120, 180, 250, 400]),
        metric=random.choice([96, 97, 98, 99]),
        gain=random.choice([18, 22, 28, 32, 35, 40, 45, 50]),
        team=random.choice([6, 8, 10, 12, 15, 20]),
        hours=random.choice([18, 24, 36, 40, 48]),
        reduced=random.choice([2, 4, 6, 8, 10]),
        value=random.choice([25, 40, 60, 85, 120, 180, 250]),
        records=random.choice([2, 5, 10, 20, 35, 50]),
    )


def generate_experience_history(
    domain: str,
    current_role: str,
    current_company: str,
    total_experience: float,
    selected_skills: list[str],
) -> str:
    blueprint = DOMAIN_BLUEPRINTS[domain]
    total_months = max(3, round(total_experience * 12))
    job_count = 1 if total_experience < 3 else 2 if total_experience < 8 else 3
    base_months = total_months // job_count
    remaining = total_months % job_count
    end_date = BASE_DATE
    entries = []

    companies = [current_company] + random.sample([c for c in COMPANIES if c != current_company], job_count - 1)
    roles = [current_role] + random.sample(blueprint["roles"], job_count - 1)

    for position in range(job_count):
        months = base_months + (1 if position < remaining else 0)
        start_date = shift_months(end_date, -months)
        end_label = "Present" if position == 0 else format_month(end_date)
        heading = f"{companies[position]} - {roles[position]}\n{format_month(start_date)} - {end_label} | {months / 12:.1f} years"
        templates = random.sample(blueprint["bullets"], k=min(4, len(blueprint["bullets"])))
        bullets = "\n".join(f"- {make_impact_bullet(item, selected_skills)}" for item in templates)
        entries.append(f"{heading}\n{bullets}")
        end_date = shift_months(start_date, -random.choice([0, 1, 2]))

    return "\n\n".join(entries)


def generate_education(domain: str, highest_education: str, total_experience: float) -> str:
    graduation_year = max(2001, 2026 - max(0, math.floor(total_experience)))
    institution = random.choice(INSTITUTIONS)
    score = random.choice(["8.1/10 CGPA", "8.6/10 CGPA", "9.0/10 CGPA", "76%", "82%", "88%"])
    primary = f"{highest_education} | {institution}\n{graduation_year - 3} - {graduation_year} | Score: {score}"
    if any(token in highest_education for token in ["MCA", "M.Sc", "MBA", "M.Com"]):
        bachelor = random.choice(EDUCATION_BY_DOMAIN[domain][:2])
        previous = f"{bachelor} | {random.choice(INSTITUTIONS)}\n{graduation_year - 6} - {graduation_year - 3} | Score: {random.choice(['7.9/10 CGPA', '74%', '81%'])}"
        return f"{primary}\n\n{previous}"
    return primary


def generate_projects(domain: str, selected_skills: list[str]) -> str:
    projects = random.sample(DOMAIN_BLUEPRINTS[domain]["projects"], k=random.randint(2, 3))
    blocks = []
    for name, tools, objective in projects:
        gain = random.choice([18, 24, 30, 35, 42, 48])
        blocks.append(
            f"{name}\n"
            f"Tools: {tools}\n"
            f"- {objective.capitalize()}.\n"
            f"- Analyzed requirements, validated source data, and documented the solution using {random.choice(selected_skills)}.\n"
            f"- Delivered measurable improvements, including a {gain}% reduction in manual effort or processing time.\n"
            f"- Presented outcomes, operating guidance, and next-step recommendations to project stakeholders."
        )
    return "\n\n".join(blocks)


def generate_certifications(domain: str, total_experience: float) -> str:
    certs = random.sample(DOMAIN_BLUEPRINTS[domain]["certifications"], k=random.randint(1, 3))
    start_year = max(2017, 2026 - max(1, math.ceil(total_experience)))
    return "\n".join(
        f"- {certification} | Issued {random.randint(start_year, 2026)} | Credential ID: HX-{random.randint(100000, 999999)}"
        for certification in certs
    )


def generate_summary(domain: str, role: str, total_experience: float, selected_skills: list[str]) -> str:
    outcome = DOMAIN_BLUEPRINTS[domain]["outcome"]
    return (
        f"{role} with {total_experience:.1f} years of experience delivering {outcome}. "
        f"Hands-on expertise in {', '.join(selected_skills[:5])}, with a record of translating business requirements "
        f"into reliable execution, measurable process improvement, and clear stakeholder reporting. "
        f"Experienced in cross-functional delivery, data validation, documentation, risk management, and continuous improvement."
    )


def grouped_skills_text(skill_groups: dict[str, list[str]]) -> str:
    return "\n".join(f"{group}: {', '.join(skills)}" for group, skills in skill_groups.items())


def generate_resume(candidate: dict, experience_history: str, skills_grouped: str) -> str:
    achievements = [
        f"- Improved a key {candidate['domain']} workflow by {random.choice([20, 25, 30, 35, 40])}% through structured analysis and automation.",
        f"- Recognized for ownership, documentation quality, and collaboration across {random.choice([3, 4, 5, 6])} cross-functional teams.",
        f"- Mentored or supported {random.choice([4, 6, 8, 10, 12])} colleagues through reviews, knowledge sessions, and reusable playbooks.",
    ]
    return (
        f"{candidate['first_name'].upper()} {candidate['last_name'].upper()}\n"
        f"{candidate['city']}, {candidate['state']}, {candidate['country']}\n"
        f"Phone: +91 {candidate['phone']} | Email: {candidate['email']}\n"
        f"LinkedIn: {candidate['linkedin_url']}\n"
        f"GitHub: {candidate['github_url'] or 'Not applicable'} | Portfolio: {candidate['portfolio_url'] or 'Not applicable'}\n\n"
        f"PROFESSIONAL SUMMARY\n{candidate['summary']}\n\n"
        f"PROFESSIONAL EXPERIENCE\n{experience_history}\n\n"
        f"TECHNICAL / PROFESSIONAL SKILLS\n{skills_grouped}\n\n"
        f"EDUCATION\n{candidate['education']}\n\n"
        f"PROJECTS & CASE STUDIES\n{candidate['projects']}\n\n"
        f"CERTIFICATIONS\n{candidate['certifications']}\n\n"
        f"SELECTED ACHIEVEMENTS\n{'\n'.join(achievements)}\n\n"
        f"ADDITIONAL INFORMATION\n"
        f"Notice Period: {candidate['notice_period']}\n"
        f"Expected Salary: INR {candidate['expected_salary']:,} per annum\n"
        f"Preferred Work Location: {candidate['city']} / Hybrid / Remote\n"
        f"Languages: English, Hindi, {random.choice(['Marathi', 'Kannada', 'Tamil', 'Telugu', 'Gujarati'])}"
    )


def generate_candidate(index: int, used_emails: set[str], used_phones: set[str]) -> dict:
    domain = random.choice(list(DOMAIN_BLUEPRINTS))
    blueprint = DOMAIN_BLUEPRINTS[domain]
    level, min_exp, max_exp = random.choice(LEVELS)
    total_experience = round(random.uniform(min_exp, max_exp), 1)
    base_role = random.choice(blueprint["roles"])
    current_role = base_role if level.lower() in base_role.lower() else f"{level} {base_role}"

    first_name = fake.first_name()
    last_name = fake.last_name()
    slug = re.sub(r"[^a-z0-9]", "", f"{first_name}.{last_name}".lower())
    email = f"{slug}.{index}@hirexdemo.com"
    while email in used_emails:
        email = f"candidate.{index}.{random.randint(100, 999)}@hirexdemo.com"
    phone = str(random.randint(7000000000, 9999999999))
    while phone in used_phones:
        phone = str(random.randint(7000000000, 9999999999))
    used_emails.add(email)
    used_phones.add(phone)

    city, state = random.choice(CITIES_STATES)
    all_skills = flatten_skills(blueprint["skills"])
    selected_skills = random.sample(all_skills, k=random.randint(8, min(12, len(all_skills))))
    current_company = random.choice(COMPANIES)
    highest_education = random.choice(EDUCATION_BY_DOMAIN[domain])

    created_at = BASE_DATE - timedelta(days=random.randint(45, 730), hours=random.randint(0, 20))
    uploaded_at = created_at + timedelta(days=random.randint(1, 30), hours=random.randint(1, 12))
    applied_at = min(BASE_DATE, uploaded_at + timedelta(days=random.randint(1, 120), hours=random.randint(1, 12)))

    summary = generate_summary(domain, current_role, total_experience, selected_skills)
    education = generate_education(domain, highest_education, total_experience)
    projects = generate_projects(domain, selected_skills)
    certifications = generate_certifications(domain, total_experience)
    experience_history = generate_experience_history(
        domain, current_role, current_company, total_experience, selected_skills
    )
    skills_grouped = grouped_skills_text(blueprint["skills"])

    candidate_id = deterministic_id("cand", index)
    candidate = {
        "candidate_id": candidate_id,
        "resume_id": deterministic_id("res", index),
        "parsed_resume_id": deterministic_id("parsed_resume", index),
        "email": email,
        "password_hash": "$pbkdf2-sha256$dummy$hirex_candidate_password_hash",
        "role": "candidate",
        "is_active": True,
        "created_at": created_at,
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "city": city,
        "state": state,
        "country": "India",
        "domain": domain,
        "level": level,
        "current_company": current_company,
        "current_role": current_role,
        "total_experience": total_experience,
        "expected_salary": salary_by_experience(total_experience),
        "notice_period": random.choice(NOTICE_PERIODS),
        "highest_education": highest_education,
        "linkedin_url": f"https://linkedin.com/in/{slug}-{index}",
        "github_url": f"https://github.com/{slug.replace('.', '')}{index}" if domain in {"Software Development", "Data Engineering", "DevOps", "Testing", "Data Analytics"} else "",
        "portfolio_url": f"https://portfolio-{slug.replace('.', '')}{index}.example.com" if domain in {"Sales and Marketing", "Data Analytics", "Software Development"} else "",
        "profile_completion_percentage": random.randint(85, 100),
        "filename": f"{candidate_id}_resume.pdf",
        "original_filename": f"{first_name}_{last_name}_Resume.pdf",
        "file_size": random.randint(180000, 950000),
        "file_type": "application/pdf",
        "storage_path": f"/storage/resumes/{candidate_id}_resume.pdf",
        "uploaded_at": uploaded_at,
        "skills": ", ".join(selected_skills),
        "total_experience_years": total_experience,
        "education": education,
        "certifications": certifications,
        "projects": projects,
        "summary": summary,
        "raw_text": "",
        "parsed_at": min(BASE_DATE, uploaded_at + timedelta(hours=random.randint(1, 48))),
        "application_status": random.choice(APPLICATION_STATUSES),
        "applied_at": applied_at,
        "resume": "",
    }
    candidate["resume"] = generate_resume(candidate, experience_history, skills_grouped)
    candidate["raw_text"] = candidate["resume"]
    return candidate


def style_workbook(writer: pd.ExcelWriter, row_count: int) -> None:
    worksheet = writer.sheets["candidates"]
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = f"A1:AQ{row_count + 1}"

    header_fill = PatternFill("solid", fgColor="285943")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(bottom=Side(style="thin", color="B7C7BE"))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    worksheet.row_dimensions[1].height = 34

    widths = {
        "candidate_id": 22, "resume_id": 22, "parsed_resume_id": 26, "email": 34,
        "password_hash": 28, "created_at": 21, "current_company": 24, "current_role": 30,
        "linkedin_url": 38, "github_url": 34, "portfolio_url": 38, "storage_path": 42,
        "uploaded_at": 21, "skills": 55, "education": 55, "certifications": 58,
        "projects": 75, "summary": 72, "raw_text": 95, "parsed_at": 21,
        "applied_at": 21, "resume": 95,
    }
    text_heavy = {"education", "certifications", "projects", "summary", "raw_text", "resume"}
    for index, column_name in enumerate(FINAL_COLUMNS, start=1):
        letter = worksheet.cell(row=1, column=index).column_letter
        worksheet.column_dimensions[letter].width = widths.get(column_name, 22)
        if column_name in text_heavy:
            for row in range(2, row_count + 2):
                worksheet.cell(row=row, column=index).alignment = Alignment(
                    vertical="top", horizontal="left", wrap_text=True
                )

    for row in range(2, row_count + 2):
        worksheet.row_dimensions[row].height = 72

    for date_column in ["created_at", "uploaded_at", "parsed_at", "applied_at"]:
        column_index = FINAL_COLUMNS.index(date_column) + 1
        for row in range(2, row_count + 2):
            worksheet.cell(row=row, column=column_index).number_format = "yyyy-mm-dd hh:mm"

    for salary_column in ["expected_salary"]:
        column_index = FINAL_COLUMNS.index(salary_column) + 1
        for row in range(2, row_count + 2):
            worksheet.cell(row=row, column=column_index).number_format = 'INR #,##0'

    table = Table(displayName="HireXCandidates", ref=f"A1:AQ{row_count + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def validate_dataframe(frame: pd.DataFrame) -> dict:
    resume_lengths = frame["resume"].str.len()
    checks = {
        "rows": len(frame),
        "columns": list(frame.columns),
        "duplicate_emails": int(frame["email"].duplicated().sum()),
        "duplicate_phones": int(frame["phone"].duplicated().sum()),
        "duplicate_candidate_ids": int(frame["candidate_id"].duplicated().sum()),
        "blank_resumes": int(frame["resume"].str.strip().eq("").sum()),
        "minimum_resume_characters": int(resume_lengths.min()),
        "average_resume_characters": round(float(resume_lengths.mean()), 2),
        "maximum_resume_characters": int(resume_lengths.max()),
        "resume_cells_over_excel_limit": int((resume_lengths > 32767).sum()),
        "resume_raw_text_mismatches": int((frame["resume"] != frame["raw_text"]).sum()),
        "invalid_application_timeline": int((frame["applied_at"] < frame["uploaded_at"]).sum()),
    }
    assert checks["rows"] == TOTAL_CANDIDATES
    assert checks["columns"] == FINAL_COLUMNS
    assert checks["duplicate_emails"] == 0
    assert checks["duplicate_phones"] == 0
    assert checks["duplicate_candidate_ids"] == 0
    assert checks["blank_resumes"] == 0
    assert checks["minimum_resume_characters"] >= 1800
    assert checks["resume_cells_over_excel_limit"] == 0
    assert checks["resume_raw_text_mismatches"] == 0
    assert checks["invalid_application_timeline"] == 0
    return checks


def main() -> None:
    used_emails: set[str] = set()
    used_phones: set[str] = set()
    candidates = [generate_candidate(i, used_emails, used_phones) for i in range(1, TOTAL_CANDIDATES + 1)]
    frame = pd.DataFrame(candidates)[FINAL_COLUMNS]
    checks = validate_dataframe(frame)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm") as writer:
        frame.to_excel(writer, sheet_name="candidates", index=False)
        style_workbook(writer, len(frame))

    print(f"Excel file generated successfully: {OUTPUT_FILE}")
    for name, value in checks.items():
        if name != "columns":
            print(f"{name}: {value}")


if __name__ == "__main__":
    main()
