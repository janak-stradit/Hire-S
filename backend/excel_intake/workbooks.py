from __future__ import annotations

import csv
from io import BytesIO, StringIO
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.formatting.rule import CellIsRule  # type: ignore
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
from openpyxl.utils import get_column_letter  # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from openpyxl.cell.cell import Cell  # type: ignore

from backend.excel_intake.contracts import CandidateIntakeRow, JobRequirementRow
from backend.taxonomy import get_taxonomy

REQUIREMENT_COLUMNS = [
    "requirement_id", "role", "department", "location", "employment_type",
    "experience_min", "experience_max", "required_skills", "preferred_skills", "education",
    "mandatory_certifications", "salary_min", "salary_max", "work_mode", "role_title_variants",
    "must_have_skills_min_experience", "preferred_industry_domain", "notice_period_max_days",
    "candidate_freshness_days", "willing_to_relocate", "source_priority", "jd_text",
    "screening_pass_score", "screening_review_score", "status", "created_by", "created_at",
]

REQUIREMENT_UPLOAD_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
REQUIREMENT_UPLOAD_CONTENT_TYPES = {
    "text/csv",
    "application/csv",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
}

PROVENANCE_COLUMNS = [
    "source_type", "verification_status", "source_reference", "agent_processing_allowed",
]

DEFAULT_JOB_REQUIREMENTS = [
    [
        "REQ-DATA-ENGINEER-001", "Data Engineer", "Data & Analytics", "Pune / Hybrid", "Full-time",
        2, 6, "Python, SQL, PostgreSQL, ETL, Data Modeling, Apache Airflow",
        "AWS, Snowflake, Spark, dbt, Docker", "Bachelor degree in Computer Science, IT, Engineering, or equivalent",
        "", 900000, 1800000, "Hybrid", "Data Pipeline Engineer, ETL Engineer, Analytics Engineer",
        "Python:2, SQL:2, Airflow:1", "Data & Analytics, SaaS, FinTech", 60, 30, True,
        "linkedin, naukri, internal_database",
        "Design, build, test, and operate reliable batch and incremental data pipelines. Develop analytics-ready data models using Python and SQL, orchestrate workflows with Apache Airflow, enforce data quality checks, optimize PostgreSQL workloads, document lineage, and collaborate with analysts and engineering teams. Experience with cloud storage, Snowflake, Spark, dbt, or Docker is preferred.",
        75, 60, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-AI-ENGINEER-001", "AI Engineer", "Artificial Intelligence", "Pune / Hybrid", "Full-time",
        3, 7, "Python, Machine Learning, NLP, LLM, FastAPI, REST APIs",
        "RAG, Vector Databases, LangChain, Docker, AWS", "Bachelor or Master degree in Computer Science, AI, Data Science, or equivalent",
        "", 1200000, 2400000, "Hybrid", "AI Engineer, LLM Engineer, GenAI Engineer",
        "Python:3, Machine Learning:2, LLM:1", "AI, SaaS, Product Engineering", 60, 30, True,
        "linkedin, naukri, github, internal_database",
        "Build production AI services using Python, machine-learning models, NLP, and large language models. Develop secure FastAPI endpoints, retrieval-augmented generation workflows, evaluation pipelines, prompt and model monitoring, and integrations with vector databases. Demonstrated deployment, testing, latency optimization, and responsible AI practices are required.",
        78, 62, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-ML-ENGINEER-001", "Machine Learning Engineer", "Artificial Intelligence", "Pune / Hybrid", "Full-time",
        3, 7, "Python, Machine Learning, scikit-learn, Pandas, NumPy, Model Deployment",
        "PyTorch, TensorFlow, MLflow, Docker, Kubernetes, AWS", "Bachelor or Master degree in Computer Science, Machine Learning, Statistics, or equivalent",
        "", 1200000, 2400000, "Hybrid", "ML Engineer, Machine Learning Developer, Applied Scientist",
        "Python:3, Machine Learning:2, Model Deployment:1", "AI, Analytics, Product Engineering", 60, 30, True,
        "linkedin, naukri, github, internal_database",
        "Develop, validate, deploy, and monitor machine-learning models. Build reproducible feature engineering and training pipelines using Python, Pandas, NumPy, and scikit-learn; expose inference services; track experiments and model versions; evaluate drift, bias, precision, recall, and latency; and partner with data engineering and product teams to move models safely into production.",
        78, 62, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-DATA-ANALYST-001", "Data Analyst", "Data & Analytics", "Pune / Hybrid", "Full-time",
        1, 5, "SQL, Excel, Power BI, Data Cleaning, Data Visualization",
        "Python, Pandas, Tableau, Snowflake, Statistics", "Bachelor degree in Computer Science, Statistics, Mathematics, Commerce, or equivalent",
        "", 600000, 1300000, "Hybrid", "Business Analyst, BI Analyst, Reporting Analyst",
        "SQL:1, Excel:1, Power BI:1", "Data & Analytics, BFSI, Retail, SaaS", 45, 30, True,
        "naukri, linkedin, internal_database",
        "Analyze multi-source business data, define and validate KPIs, perform data cleaning and reconciliation, write efficient SQL, and build clear Power BI or Tableau dashboards. Translate stakeholder questions into repeatable analysis, explain trends and anomalies, document assumptions, and deliver decision-ready insights with attention to data quality.",
        72, 58, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-BACKEND-ENGINEER-001", "Backend Engineer", "Engineering", "Pune / Hybrid", "Full-time",
        2, 6, "Python, FastAPI, PostgreSQL, SQLAlchemy, REST APIs",
        "Redis, Docker, AWS, Celery, Microservices", "Bachelor degree in Computer Science, IT, Engineering, or equivalent",
        "", 900000, 1900000, "Hybrid", "Python Backend Engineer, API Engineer, Software Engineer Backend",
        "Python:2, FastAPI:1, PostgreSQL:1", "SaaS, FinTech, Product Engineering", 60, 30, True,
        "linkedin, naukri, github, internal_database",
        "Design and maintain secure backend services and REST APIs using Python, FastAPI, PostgreSQL, and SQLAlchemy. Implement authentication, validation, database migrations, automated tests, observability, background processing, and performance improvements while collaborating with frontend and product teams.",
        75, 60, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-FRONTEND-ENGINEER-001", "Frontend Engineer", "Engineering", "Pune / Hybrid", "Full-time",
        2, 6, "JavaScript, TypeScript, React, HTML, CSS, REST APIs",
        "Redux, Tailwind CSS, Testing Library, Vite, Accessibility", "Bachelor degree in Computer Science, IT, Engineering, or equivalent",
        "", 800000, 1700000, "Hybrid", "React Developer, UI Engineer, Frontend Developer",
        "React:2, TypeScript:1, REST APIs:1", "SaaS, Product Engineering, E-commerce", 60, 30, True,
        "linkedin, naukri, github, internal_database",
        "Build responsive, accessible, and maintainable web applications with React and TypeScript. Convert product requirements into reusable components, integrate REST APIs, manage client and server state, test critical workflows, improve performance, and collaborate closely with design and backend engineering.",
        74, 59, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-DEVOPS-ENGINEER-001", "DevOps Engineer", "Platform Engineering", "Pune / Hybrid", "Full-time",
        3, 7, "Linux, Docker, CI/CD, AWS, Git, Monitoring",
        "Kubernetes, Terraform, Python, Prometheus, Grafana", "Bachelor degree in Computer Science, IT, Engineering, or equivalent",
        "", 1000000, 2200000, "Hybrid", "Platform Engineer, Cloud Engineer, Site Reliability Engineer",
        "Linux:2, Docker:2, CI/CD:1, AWS:1", "Cloud, SaaS, Platform Engineering", 60, 30, True,
        "linkedin, naukri, github, internal_database",
        "Build and operate secure CI/CD pipelines, containerized environments, cloud infrastructure, monitoring, alerting, backup, and incident-response practices. Automate repeatable operations, improve deployment reliability, manage secrets and access, document runbooks, and partner with development teams on production readiness.",
        76, 60, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-QA-AUTOMATION-001", "QA Automation Engineer", "Quality Engineering", "Pune / Hybrid", "Full-time",
        2, 6, "Test Automation, API Testing, Selenium, Python, SQL",
        "Playwright, Pytest, Postman, CI/CD, Performance Testing", "Bachelor degree in Computer Science, IT, Engineering, or equivalent",
        "", 700000, 1500000, "Hybrid", "SDET, Automation Tester, Test Automation Engineer",
        "Test Automation:2, API Testing:1, Selenium:1", "Quality Engineering, SaaS, Product Engineering", 60, 30, True,
        "naukri, linkedin, github, internal_database",
        "Create risk-based test plans and maintain reliable UI, API, integration, and regression automation. Investigate defects, validate data with SQL, integrate tests into CI/CD, report quality metrics, improve test coverage, and collaborate with engineers and product owners on acceptance criteria and release readiness.",
        73, 58, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-IAM-PAM-ENGINEER-001", "IAM / PAM Engineer", "Cybersecurity", "Pune / Hybrid", "Full-time",
        2, 7, "Identity Access Management, Privileged Access Management, Active Directory, CyberArk",
        "BeyondTrust, PowerShell, Azure AD, SAML, OAuth", "Bachelor degree in Computer Science, IT, Cybersecurity, or equivalent",
        "", 900000, 2000000, "Hybrid", "IAM Engineer, PAM Analyst, Identity Security Engineer",
        "Identity Access Management:2, Privileged Access Management:1, Active Directory:1",
        "Cybersecurity, IT Operations, Enterprise Security", 60, 30, True,
        "naukri, linkedin, internal_database",
        "Implement and support enterprise IAM and PAM controls, privileged account onboarding, safes and platform management, password rotation, session monitoring, access reviews, incident handling, compliance reporting, and integrations with Active Directory and cloud identity platforms. Maintain operational documentation and troubleshoot authentication issues.",
        75, 60, "draft", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
    [
        "REQ-IT-SUPPORT-ENGINEER-001", "IT Support Engineer", "IT Support", "Pune / On-site", "Full-time",
        1, 6, "IT Support, Technical Troubleshooting, Active Directory, Ticketing Systems, Hardware Support, Network Basics",
        "O365, ITSM, Windows Server, ServiceNow, TCP/IP, ITIL, Endpoint Security, Jira Service Desk",
        "Bachelor degree or diploma in Computer Science, IT, Electronics, or equivalent", "", 450000, 1200000,
        "On-site", "Desktop Support Engineer, Technical Support Engineer, Service Desk Engineer",
        "IT Support:1, Technical Troubleshooting:1, Active Directory:1",
        "IT Operations, Managed Services, Enterprise Support", 45, 30, True,
        "naukri, linkedin, internal_database",
        "Provide reliable end-user and infrastructure support across Windows endpoints, hardware, software, identity, and basic networking. Diagnose incidents, manage tickets and SLAs, support Active Directory accounts and access, document resolutions, escalate complex problems, maintain asset and knowledge records, and communicate clearly with users. Experience with O365, ServiceNow, ITSM, Windows Server, TCP/IP, ITIL, or endpoint security is preferred.",
        72, 58, "active", "HR / Talent Lead", datetime.now(UTC).replace(tzinfo=None),
    ],
]


class WorkbookValidationError(ValueError):
    pass


def _records(path: Path, preferred_sheet: str) -> list[dict[str, Any]]:
    if not path.exists():
        raise WorkbookValidationError(f"Workbook not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    _active = workbook[preferred_sheet] if preferred_sheet in workbook.sheetnames else workbook.active
    assert _active is not None, f"Workbook has no active sheet: {path}"
    sheet = _active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration as exc:
        raise WorkbookValidationError(f"Workbook is empty: {path}") from exc
    if len(headers) != len(set(headers)):
        raise WorkbookValidationError(f"Workbook contains duplicate headers: {path}")
    return [dict(zip(headers, row, strict=True)) for row in rows if any(value is not None for value in row)]


def _list(value: Any) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).replace(";", ",").split(",") if part.strip()]


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _slug(value: str) -> str:
    slug = "".join(ch if ch.isalnum() else "-" for ch in value.upper()).strip("-")
    while "--" in slug:
        slug = slug.replace("--", "-")
    return slug or "JOB"


def _next_requirement_id(role: str, used_ids: set[str]) -> str:
    base = f"REQ-{_slug(role)}"
    index = 1
    candidate = f"{base}-{index:03d}"
    while candidate in used_ids:
        index += 1
        candidate = f"{base}-{index:03d}"
    used_ids.add(candidate)
    return candidate


def _normalize_requirement_payload(row: dict[str, Any], used_ids: set[str]) -> dict[str, Any]:
    payload = {column: row.get(column) for column in REQUIREMENT_COLUMNS}
    role = str(payload.get("role") or "").strip()
    payload["requirement_id"] = str(payload.get("requirement_id") or "").strip() or _next_requirement_id(role, used_ids)
    used_ids.add(payload["requirement_id"])
    payload["employment_type"] = payload.get("employment_type") or "Full-time"
    payload["work_mode"] = payload.get("work_mode") or "Hybrid"
    payload["candidate_freshness_days"] = payload.get("candidate_freshness_days") or 30
    payload["screening_pass_score"] = payload.get("screening_pass_score") or 75
    payload["screening_review_score"] = payload.get("screening_review_score") or 60
    payload["status"] = payload.get("status") or "draft"
    for key in (
        "required_skills", "preferred_skills", "mandatory_certifications",
        "role_title_variants", "source_priority",
    ):
        payload[key] = _list(payload.get(key))
    for key in ("must_have_skills_min_experience", "preferred_industry_domain", "education", "jd_text"):
        payload[key] = str(payload.get(key) or "").strip()
    if payload.get("willing_to_relocate") not in {None, ""}:
        payload["willing_to_relocate"] = _bool(payload.get("willing_to_relocate"))
    return payload


def list_requirements(path: Path) -> list[JobRequirementRow]:
    rows = _records(path, "Job_Requirements")
    required_columns = {
        "requirement_id", "role", "department", "location", "employment_type",
        "experience_min", "experience_max", "required_skills", "preferred_skills",
        "education", "mandatory_certifications", "salary_min", "salary_max",
        "jd_text", "screening_pass_score", "screening_review_score", "status",
    }
    missing = required_columns - set(rows[0] if rows else {})
    if missing:
        raise WorkbookValidationError(f"Missing job requirement columns: {', '.join(sorted(missing))}")
    requirements = []
    for row in rows:
        payload = dict(row)
        payload["required_skills"] = _list(payload.get("required_skills"))
        payload["preferred_skills"] = _list(payload.get("preferred_skills"))
        payload["mandatory_certifications"] = _list(payload.get("mandatory_certifications"))
        payload["role_title_variants"] = _list(payload.get("role_title_variants"))
        payload["source_priority"] = _list(payload.get("source_priority"))
        for key in ("must_have_skills_min_experience", "preferred_industry_domain", "work_mode"):
            if payload.get(key) is None:
                payload[key] = "" if key != "work_mode" else "Hybrid"
        if payload.get("candidate_freshness_days") in {None, ""}:
            payload["candidate_freshness_days"] = 30
        if payload.get("willing_to_relocate") not in {None, ""}:
            payload["willing_to_relocate"] = _bool(payload.get("willing_to_relocate"))
        requirements.append(JobRequirementRow.model_validate(payload))
    return requirements


def read_requirement(path: Path, requirement_id: str | None = None) -> JobRequirementRow:
    requirements = list_requirements(path)
    if requirement_id:
        selected = [row for row in requirements if row.requirement_id == requirement_id]
        if len(selected) != 1:
            raise WorkbookValidationError(f"Job requirement not found: {requirement_id}")
        return selected[0]
    active = [row for row in requirements if row.status.strip().lower() in {"active", "open"}]
    if len(active) != 1:
        raise WorkbookValidationError("Exactly one active/open job requirement row is required")
    return active[0]


def read_active_requirement(path: Path) -> JobRequirementRow:
    return read_requirement(path)


def upsert_requirement(path: Path, requirement: JobRequirementRow, created_by: str) -> None:
    workbook = load_workbook(path)
    _active = workbook["Job_Requirements"] if "Job_Requirements" in workbook.sheetnames else workbook.active
    assert _active is not None, "Workbook has no active sheet"
    sheet = _active
    if not isinstance(sheet, Worksheet):
        raise WorkbookValidationError("Valid worksheet not found")
    headers = _ensure_requirement_columns(sheet)
    row_index = next(
        (row for row in range(2, sheet.max_row + 1) if sheet.cell(row, 1).value == requirement.requirement_id),
        sheet.max_row + 1,
    )
    values = requirement.model_dump()
    for key in ("required_skills", "preferred_skills", "mandatory_certifications", "role_title_variants", "source_priority"):
        values[key] = ", ".join(values[key])
    values.update(
        created_by=created_by,
        created_at=datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
    )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row_index, column, values.get(header))
    created_at_column = headers.index("created_at") + 1
    sheet.cell(row_index, created_at_column).number_format = "@"
    if sheet.tables:
        tbl: Table = next(iter(sheet.tables.values()))  # type: ignore[assignment]
        tbl.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    workbook.save(path)


def import_requirements(path: Path, filename: str, content: bytes, created_by: str) -> dict[str, Any]:
    """Import many JD rows from a CSV/XLSX file into the JD workbook.

    Required input columns: role, department, location, required_skills, jd_text.
    All other JD columns are optional and receive the same defaults as the
    single-create form. Blank requirement_id values are generated automatically.
    """
    suffix = Path(filename).suffix.lower()
    if suffix not in REQUIREMENT_UPLOAD_EXTENSIONS:
        raise WorkbookValidationError("Only CSV and Excel files are allowed for bulk JD upload (.csv, .xlsx, .xlsm)")
    rows = _uploaded_requirement_rows(filename, content)
    if not rows:
        raise WorkbookValidationError("Bulk JD upload file does not contain any data rows")
    workbook = load_workbook(path)
    _active = workbook["Job_Requirements"] if "Job_Requirements" in workbook.sheetnames else workbook.active
    assert _active is not None, "Workbook has no active sheet"
    sheet = _active
    if not isinstance(sheet, Worksheet):
        raise WorkbookValidationError("Valid worksheet not found")
    headers = _ensure_requirement_columns(sheet)
    existing_ids = {
        str(sheet.cell(row, 1).value or "").strip()
        for row in range(2, sheet.max_row + 1)
        if str(sheet.cell(row, 1).value or "").strip()
    }
    used_ids = set(existing_ids)
    existing_row_by_id = {
        str(sheet.cell(row, 1).value or "").strip(): row
        for row in range(2, sheet.max_row + 1)
        if str(sheet.cell(row, 1).value or "").strip()
    }
    status_column = headers.index("status") + 1
    errors: list[dict[str, Any]] = []
    imported: list[str] = []
    created = 0
    updated = 0
    processed = 0
    for index, row in enumerate(rows, start=2):
        if not any(value not in {None, ""} for value in row.values()):
            continue
        processed += 1
        try:
            payload = _normalize_requirement_payload(row, used_ids)
            requirement = JobRequirementRow.model_validate(payload)
        except Exception as exc:
            errors.append({"row": index, "error": str(exc)})
            continue
        values = requirement.model_dump()
        for key in ("required_skills", "preferred_skills", "mandatory_certifications", "role_title_variants", "source_priority"):
            values[key] = ", ".join(values[key])
        values.update(
            created_by=created_by,
            created_at=datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z"),
        )
        if requirement.status.strip().lower() == "active":
            for row_number in range(2, sheet.max_row + 1):
                if str(sheet.cell(row_number, status_column).value or "").strip().lower() == "active":
                    sheet.cell(row_number, status_column, "inactive")
        target_row = existing_row_by_id.get(requirement.requirement_id)
        if target_row:
            updated += 1
        else:
            target_row = sheet.max_row + 1
            existing_row_by_id[requirement.requirement_id] = target_row
            created += 1
        for column, header in enumerate(headers, start=1):
            sheet.cell(target_row, column, values.get(header))
        created_at_column = headers.index("created_at") + 1
        sheet.cell(target_row, created_at_column).number_format = "@"
        imported.append(requirement.requirement_id)
    if sheet.tables:
        tbl: Table = next(iter(sheet.tables.values()))  # type: ignore[assignment]
        tbl.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"
    workbook.save(path)
    return {
        "filename": filename,
        "rows_read": len(rows),
        "rows_processed": processed,
        "created": created,
        "updated": updated,
        "failed": len(errors),
        "imported_requirement_ids": imported,
        "errors": errors[:100],
    }


def _uploaded_requirement_rows(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        return [
            {str(key).strip(): value for key, value in row.items() if key is not None}
            for row in csv.DictReader(StringIO(text))
        ]
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        _active = workbook["Job_Requirements"] if "Job_Requirements" in workbook.sheetnames else workbook.active
        assert _active is not None, "Workbook has no active sheet"
        sheet = _active
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        except StopIteration:
            workbook.close()
            return []
        parsed = [
            dict(zip(headers, row, strict=True))
            for row in rows
            if any(value is not None for value in row)
        ]
        workbook.close()
        return parsed
    raise WorkbookValidationError("Bulk JD upload must be a .csv, .xlsx, or .xlsm file")


def update_requirement_status(path: Path, requirement_id: str, status: str) -> JobRequirementRow:
    workbook = load_workbook(path)
    _active = workbook["Job_Requirements"] if "Job_Requirements" in workbook.sheetnames else workbook.active
    assert _active is not None, "Workbook has no active sheet"
    sheet: Worksheet = _active  # type: ignore[assignment]
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    status_column = headers.index("status") + 1
    target_row = None
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value or "").strip() == requirement_id:
            target_row = row
        if status == "active" and str(sheet.cell(row, status_column).value or "").strip().lower() == "active":
            sheet.cell(row, status_column, "inactive")
    if target_row is None:
        workbook.close()
        raise WorkbookValidationError(f"Job requirement not found: {requirement_id}")
    sheet.cell(target_row, status_column, status)
    workbook.save(path)
    return read_requirement(path, requirement_id)


def read_candidates(path: Path, include_synthetic: bool = False) -> tuple[list[CandidateIntakeRow], int]:
    rows = _records(path, "Candidates")
    required = {"email", *PROVENANCE_COLUMNS}
    missing = required - set(rows[0] if rows else {})
    if missing:
        raise WorkbookValidationError(f"Missing candidate columns: {', '.join(sorted(missing))}")
    candidates: list[CandidateIntakeRow] = []
    seen_emails: set[str] = set()
    skipped = 0
    for row_number, row in enumerate(rows, start=2):
        source_type = str(row.get("source_type") or "").strip().lower()
        allowed = _bool(row.get("agent_processing_allowed"))
        simulation_allowed = source_type == "synthetic" and include_synthetic
        if not allowed and not simulation_allowed:
            skipped += 1
            continue
        email = str(row.get("email") or "").strip().lower()
        if not email or email in seen_emails:
            raise WorkbookValidationError(f"Missing or duplicate email at candidate row {row_number}: {email}")
        seen_emails.add(email)
        payload = {
            key: row.get(key)
            for key in CandidateIntakeRow.model_fields
            if row.get(key) is not None
        }
        payload.update(email=email, source_type=source_type, agent_processing_allowed=allowed)
        candidates.append(CandidateIntakeRow.model_validate(payload))
    return candidates, skipped


def create_requirement_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    _active = workbook.active
    if not isinstance(_active, Worksheet):
        raise ValueError("Worksheet creation failed")
    sheet: Worksheet = _active
    sheet.title = "Job_Requirements"
    sheet.append(REQUIREMENT_COLUMNS)
    taxonomy = get_taxonomy()
    for source_requirement in DEFAULT_JOB_REQUIREMENTS:
        requirement = list(source_requirement)
        requirement[26] = datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")
        primary, secondary = taxonomy.role_keywords(str(requirement[1]))
        required = _list(requirement[7])
        enriched_preferred = list(dict.fromkeys([
            *_list(requirement[8]),
            *(skill for skill in [*primary, *secondary] if skill not in required),
        ]))
        requirement[8] = ", ".join(enriched_preferred[:24])
        sheet.append(requirement)
    end_column = get_column_letter(len(REQUIREMENT_COLUMNS))
    _style_table(sheet, f"A1:{end_column}{sheet.max_row}", "JobRequirements")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{end_column}{sheet.max_row}"
    sheet.row_dimensions[1].height = 28
    widths = {"A": 16, "B": 24, "C": 18, "D": 20, "E": 16, "H": 36, "I": 30, "J": 38, "K": 28, "N": 16, "O": 32, "P": 30, "Q": 26, "U": 28, "V": 60, "Y": 14, "Z": 24, "AA": 20}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        last_cell: Cell | None = None
        for cell in row:
            if isinstance(cell, Cell):
                cell.alignment = cast(Any, Alignment(vertical="top", wrap_text=cell.column in {8, 9, 10, 11, 14}))
                last_cell = cell
        if last_cell is not None:
            sheet.row_dimensions[last_cell.row].height = 90
    for cell in sheet["AA"][1:]:
        cell.number_format = "@"
    status_validation = DataValidation(type="list", formula1='"draft,active,inactive"')
    sheet.add_data_validation(cast(Any, status_validation))
    status_validation.add("Y2:Y500")
    workbook.save(path)


def _ensure_requirement_columns(sheet: Worksheet) -> list[str]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    for column in REQUIREMENT_COLUMNS:
        if column not in headers:
            headers.append(column)
            sheet.cell(1, len(headers), column)
    return headers


def export_results(path: Path, rows: list[dict[str, Any]], title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = [
        "batch_id", "requirement_id", "job_id", "application_id", "candidate_id", "email",
        "candidate_name", "current_role", "experience_years", "decision", "queue_target",
        "final_score", "skill_score", "experience_score", "education_score",
        "certification_score", "keyword_score", "explanation", "source_type",
        "verification_status", "source_reference", "evaluated_at",
    ]
    workbook = Workbook()
    sheet = workbook.active
    if not isinstance(sheet, Worksheet):
        raise ValueError("Worksheet creation failed")
    cast(Any, sheet).title = title[:31]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    end_row = max(sheet.max_row, 2)
    _style_table(sheet, f"A1:V{end_row}", "ScreeningResults")
    cast(Any, sheet).freeze_panes = "A2"
    cast(Any, sheet.auto_filter).ref = f"A1:V{end_row}"
    for column in ("L", "M", "N", "O", "P", "Q"):
        for cell in sheet[column][1:]:
            cast(Any, cell).number_format = "0.00"
    sheet.conditional_formatting.add(
        f"L2:L{end_row}", CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=PatternFill("solid", fgColor="DCFCE7"))
    )
    widths = {"A": 38, "B": 18, "C": 38, "D": 38, "E": 24, "F": 34, "G": 24, "H": 26, "J": 12, "K": 20, "L": 14, "R": 60, "S": 18, "T": 22, "U": 44, "V": 22}
    for column, width in widths.items():
        cast(Any, sheet.column_dimensions[column]).width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell, Cell):
                cell.alignment = cast(Any, Alignment(vertical="top", wrap_text=cell.column in {18, 21}))
    workbook.save(path)


def _style_table(sheet: Worksheet, reference: str, name: str) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    sheet.add_table(cast(Any, table))
    for cell in sheet[1]:
        if isinstance(cell, Cell):
            cell.fill = cast(Any, PatternFill("solid", fgColor="0F766E"))
            cell.font = cast(Any, Font(color="FFFFFF", bold=True))
            cell.alignment = cast(Any, Alignment(vertical="center", wrap_text=True))
